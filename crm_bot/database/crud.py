from sqlalchemy import select, join, update, func
from sqlalchemy.orm import joinedload, selectinload
from sqlalchemy.exc import SQLAlchemyError
from datetime import datetime
from openpyxl import Workbook
from io import BytesIO
from database.models import Seller, Client, Order, Consumptions
from database.utils import async_session
from .database import AsyncSessionLocal

async def get_seller_by_passport(passport_serial: str):
    try:
        async with async_session() as session:
            result = await session.execute(
                select(Seller).where(Seller.passport_serial == passport_serial)
            )
            return result.scalar_one_or_none()
    except SQLAlchemyError as e:
        return None

async def add_seller_to_db(data: dict):
    async with async_session() as session:
        try:
            seller = Seller(
                full_name=data['full_name'],
                phone=data['phone'],
                passport_serial=data['passport_serial'],
                salary_of_seller=data.get('salary_of_seller'),
                started_job_at=data['started_job_at'],
            )
            session.add(seller)
            await session.commit()
            await session.refresh(seller)
            return seller
        except Exception as e:
            await session.rollback()
            raise e

async def add_client_to_db(client_data: dict):
    async with AsyncSessionLocal() as session:
        try:
            # Ensure required location fields are present
            if 'latitude' not in client_data or 'longitude' not in client_data:
                raise ValueError("Location coordinates are required")
                
            client = Client(**client_data)
            session.add(client)
            await session.commit()
            await session.refresh(client)
            return client
        except Exception as e:
            await session.rollback()
            raise e

async def get_client_by_passport(passport: str):
    try:
        async with async_session() as session:
            result = await session.execute(
                select(Client).where(Client.passport_serial == passport)
            )
            return result.scalars().first()
    except SQLAlchemyError as e:
        return None

async def create_order(order_data: dict):
    async with AsyncSessionLocal() as session:
        try:
            # Set default status if not provided
            order_data.setdefault('order_status', 'Ochiq')
            order = Order(**order_data)
            session.add(order)
            await session.commit()
            await session.refresh(order)
            return order
        except Exception as e:
            await session.rollback()
            raise e

async def get_all_orders_with_details():
    async with async_session() as session:
        try:
            query = (
                select(
                    Order.id.label("order_id"),
                    Order.order_status,
                    Order.created_at,
                    Order.item_count,
                    Order.sum_of_item,
                    Order.every_month_should_pay,
                    Order.prepaid,
                    Order.total_paid,
                    Order.remaining_amount,
                    Client.full_name.label("client_name"),
                    Client.phone.label("client_phone"),
                    Client.passport_serial.label("client_passport"),
                    Client.latitude.label("client_latitude"),  # Added
                    Client.longitude.label("client_longitude"),  # Added
                    Seller.full_name.label("seller_name")
                )
                .select_from(
                    join(Order, Client, Order.client_id == Client.id)
                    .join(Seller, Order.seller_id == Seller.id)
                )
                .order_by(Order.created_at.desc())
            )
            result = await session.execute(query)
            return result.all()
        except Exception as e:
            return None


async def generate_orders_excel():
    orders = await get_all_orders_with_details()
    if not orders:
        return None
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    
    headers = [
        "Buyurtma ID", "Status", "Sana", "Mijoz", "Telefon",
        "Passport", "Joylashuv", "Sotuvchi", "Mahsulot Soni",
        "Umumiy Summa", "Oylik To'lov", "Oldindan To'lov","Ja'mi to'langan summa","Qoldiq"
    ]
    ws.append(headers)
    
    for order in orders:
        ws.append([
            order.order_id,
            order.order_status,
            order.created_at.strftime("%Y-%m-%d %H:%M"),
            order.client_name,
            order.client_phone,
            order.client_passport,
            f"https://maps.google.com/?q={order.client_latitude},{order.client_longitude}",  # Location format
            order.seller_name,
            order.item_count,
            order.sum_of_item,
            order.every_month_should_pay,
            order.prepaid,
            order.total_paid,
            order.remaining_amount
        ])
    
    # Header styling
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    
    # Auto-adjust columns
    for column in ws.columns:
        max_length = max(
            len(str(cell.value)) for cell in column
        )
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


async def get_all_sellers_with_details():
    async with async_session() as session:
        try:
            result = await session.execute(
                select(
                    Seller.id.label("seller_id"),
                    Seller.full_name.label("full_name"),
                    Seller.phone,
                    Seller.passport_serial,
                    Seller.started_job_at,
                    Seller.order_counter,
                    Seller.salary_of_seller.label("salary_of_seller")
                )
                .order_by(Seller.full_name)
            )
            return result.all()
        except Exception as e:
            return None

async def generate_sellers_excel():
    sellers = await get_all_sellers_with_details()
    if not sellers:
        return None
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sellers"
    
    headers = [
        "ID", "To'liq Ism", "Telefon", "Passport Seriya", 
        "Ish Boshlagan Sana", "Sotgan Mahsulotlar Soni", "Maosh"
    ]
    ws.append(headers)
    
    for seller in sellers:
        ws.append([
            seller.seller_id,
            seller.full_name,
            seller.phone,
            seller.passport_serial,
            seller.started_job_at.strftime("%Y-%m-%d") if seller.started_job_at else "",
            seller.order_counter,
            seller.salary_of_seller
        ])
    
    # Header styling
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    
    # Auto-adjust columns
    for column in ws.columns:
        max_length = max(
            len(str(cell.value)) for cell in column
        )
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Format date columns (column E)
    for cell in ws['E'][1:]:  # Skip header row
        cell.number_format = 'YYYY-MM-DD'
    
    # Format salary column (column G) as currency
    for cell in ws['G'][1:]:  # Skip header row
        cell.number_format = '#,##0.00'
    
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

async def add_monthly_payment(order_id: int, amount: int):
    """Добавление ежемесячного платежа к заказу"""
    async with AsyncSessionLocal() as session:
        try:
            order = await session.get(Order, order_id)
            if not order:
                return False
                
            order.total_paid += amount
            order.remaining_amount = max(0, order.sum_of_item - order.total_paid)
            
            await session.commit()
            return True
        except Exception as e:
            await session.rollback()
            raise e

async def update_order(order_id: int, update_data: dict):
    """Обновление данных заказа с полной проверкой полей"""
    async with AsyncSessionLocal() as session:
        try:
            # Allowed fields - removed item_returned since it's now part of order_status
            allowed_fields = {
                'item_count', 'sum_of_item', 'every_month_should_pay',
                'prepaid', 'order_status'
            }
            
            db_update_data = {
                field: value 
                for field, value in update_data.items() 
                if field in allowed_fields
            }
            
            if not db_update_data:
                raise ValueError("Yangilanish uchun hech qanday maydon kiritilmadi")
            
            # Автоматический пересчет remaining_amount
            if {'sum_of_item', 'prepaid'} & db_update_data.keys():
                order = await session.get(Order, order_id)
                if order:
                    new_sum = db_update_data.get('sum_of_item', order.sum_of_item)
                    new_prepaid = db_update_data.get('prepaid', order.prepaid)
                    db_update_data['remaining_amount'] = max(0, new_sum - new_prepaid)
            
            await session.execute(
                update(Order)
                .where(Order.id == order_id)
                .values(**db_update_data)
            )
            await session.commit()
            return True
        except Exception as e:
            await session.rollback()
            raise Exception(f"Xatolik yuz berdi: {str(e)}")

async def get_order_by_id_with_details(order_id: int):
    async with async_session() as session:
        result = await session.execute(
            select(Order)
            .options(joinedload(Order.client))  # Eager load client
            .where(Order.id == order_id)
        )
        return result.scalars().first()

async def delete_order(order_id: int):
    """Удаление заказа"""
    async with AsyncSessionLocal() as session:
        try:
            order = await session.get(Order, order_id)
            if not order:
                return False
                
            await session.delete(order)
            await session.commit()
            return True
        except Exception as e:
            await session.rollback()
            return False


async def update_seller(seller_id: int, update_data: dict):
    """Обновление данных продавца с преобразованием типов данных"""
    async with async_session() as session:
        try:
            # Полный маппинг между именами полей
            field_mapping = {
                'full_name': 'full_name',
                'name': 'full_name',
                'phone': 'phone',
                'passport': 'passport_serial',
                'passport_serial': 'passport_serial',
                'salary': 'salary_of_seller',
                'salary_of_seller': 'salary_of_seller',
                'date': 'started_job_at',
                'start_date': 'started_job_at',
                'started_job_at': 'started_job_at'
            }
            
            # Проверяем и преобразуем данные
            db_update_data = {}
            
            for field, value in update_data.items():
                if field in field_mapping:
                    db_field = field_mapping[field]
                    
                    # Специальная обработка для полей даты
                    if db_field == 'started_job_at' and isinstance(value, str):
                        try:
                            db_update_data[db_field] = datetime.strptime(value, '%Y-%m-%d').date()
                        except ValueError:
                            raise ValueError("Noto'g'ri sana formati. To'g'ri format: YYYY-MM-DD")
                    else:
                        db_update_data[db_field] = value
            
            if not db_update_data:
                raise ValueError("Yangilanish uchun hech qanday maydon kiritilmadi")
            
            await session.execute(
                update(Seller)
                .where(Seller.id == seller_id)
                .values(**db_update_data)
            )
            await session.commit()
            return True
            
        except ValueError as e:
            await session.rollback()
            raise ValueError(f"Ma'lumotlar formati noto'g'ri: {str(e)}")
        except Exception as e:
            await session.rollback()
            raise Exception(f"Xatolik yuz berdi: {str(e)}")

async def get_seller_by_id_or_passport(seller_id: int = None, passport_serial: str = None):
    async with async_session() as session:
        query = select(Seller)
        if seller_id:
            query = query.where(Seller.id == seller_id)
        elif passport_serial:
            query = query.where(Seller.passport_serial == passport_serial)
        else:
            return None
            
        result = await session.execute(query)
        return result.scalars().first()

async def delete_seller(seller_id: int):
    async with async_session() as session:
        try:
            seller = await session.get(Seller, seller_id)
            if seller:
                await session.delete(seller)
                await session.commit()
                return True
            return False
        except Exception as e:
            await session.rollback()
            return False

async def get_consumption_by_id(consumption_id: int):
    """Get a single consumption record by ID"""
    async with async_session() as session:
        result = await session.execute(
            select(Consumptions)
            .where(Consumptions.id == consumption_id)
        )
        return result.scalars().first()

async def get_consumptions_by_owner(owner: str):
    """Get all consumptions for a specific owner"""
    async with async_session() as session:
        result = await session.execute(
            select(Consumptions)
            .where(Consumptions.consumption_owner == owner)
            .order_by(Consumptions.created_at.desc())
        )
        return result.scalars().all()

async def create_consumption(owner: str, amount: float, description: str):
    """Create a new consumption record"""
    async with async_session() as session:
        try:
            consumption = Consumptions(
                consumption_owner=owner,
                amount=amount,
                description=description
            )
            session.add(consumption)
            await session.commit()
            await session.refresh(consumption)
            return consumption
        except Exception as e:
            await session.rollback()
            raise Exception(f"Error creating consumption: {str(e)}")

async def update_consumption(consumption_id: int, update_data: dict):
    """Update consumption record with data validation"""
    async with async_session() as session:
        try:
            # Field mapping with validation
            valid_fields = {
                'amount': (float, "Raqam bo'lishi kerak"),
                'description': (str, "Matn bo'lishi kerak"),
                'owner': ('consumption_owner', str, "Egasi nomi bo'lishi kerak")
            }
            
            db_update_data = {}
            
            for field, value in update_data.items():
                if field in valid_fields:
                    if field == 'owner':
                        db_field, field_type, error_msg = valid_fields[field]
                        if not isinstance(value, field_type):
                            raise ValueError(f"{field}: {error_msg}")
                        db_update_data[db_field] = value
                    else:
                        field_type, error_msg = valid_fields[field]
                        try:
                            db_update_data[field] = field_type(value)
                        except (ValueError, TypeError):
                            raise ValueError(f"{field}: {error_msg}")
            
            if not db_update_data:
                raise ValueError("Yangilanish uchun hech qanday maydon kiritilmadi")
            
            await session.execute(
                update(Consumptions)
                .where(Consumptions.id == consumption_id)
                .values(**db_update_data)
            )
            await session.commit()
            return True
            
        except ValueError as e:
            await session.rollback()
            raise ValueError(f"Validation error: {str(e)}")
        except Exception as e:
            await session.rollback()
            raise Exception(f"Database error: {str(e)}")

async def get_all_consumptions():
    """Get all consumption records with details"""
    async with async_session() as session:
        try:
            result = await session.execute(
                select(
                    Consumptions.id.label("consumption_id"),
                    Consumptions.consumption_owner.label("owner"),
                    Consumptions.amount,
                    Consumptions.description,
                    Consumptions.created_at
                )
                .order_by(Consumptions.created_at.desc())
            )
            return result.all()
        except Exception as e:
            return None

async def generate_consumptions_excel(owner: str = None):
    """Generate Excel report for consumptions (optionally filtered by owner)"""
    if owner:
        consumptions = await get_consumptions_by_owner(owner)
    else:
        consumptions = await get_all_consumptions()
    
    if not consumptions:
        return None
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Consumptions"
    
    headers = [
        "ID", "Egasi", "Summa", 
        "Tavsifi", "Sana"
    ]
    ws.append(headers)
    
    for cons in consumptions:
        ws.append([
            cons.consumption_id if hasattr(cons, 'consumption_id') else cons.id,
            cons.owner if hasattr(cons, 'owner') else cons.consumption_owner,
            cons.amount,
            cons.description,
            cons.created_at.strftime("%Y-%m-%d %H:%M") if cons.created_at else ""
        ])
    
    # Header styling
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    
    # Auto-adjust columns
    for column in ws.columns:
        max_length = max(
            len(str(cell.value)) for cell in column
        )
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

async def get_total_consumptions_by_owner():
    """Get total consumption amounts grouped by owner"""
    async with async_session() as session:
        try:
            result = await session.execute(
                select(
                    Consumptions.consumption_owner,
                    func.sum(Consumptions.amount).label("total_amount")
                )
                .group_by(Consumptions.consumption_owner)
                .order_by(Consumptions.consumption_owner)
            )
            return result.all()
        except Exception as e:
            return None

async def delete_consumption(consumption_id: int):
    """Delete a consumption record by ID"""
    async with async_session() as session:
        try:
            consumption = await session.get(Consumptions, consumption_id)
            
            if consumption:
                await session.delete(consumption)
                await session.commit()
                return True
            return False  # Если расход не найден
            
        except Exception as e:
            await session.rollback()
            return False  # Возвращаем False при ошибке
